import datetime

from operator import attrgetter

import copy

from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.styles import Alignment
import openpyxl as xl

from modules.crm_actions import CRM

from models.Animator import Animator
from models.Game import Game
from models.GameAnim import GameAnim
from models.Helper import Helper
from models.LuftTime import LuftTime
from models.Quest import Quest
from models.Zone import Zone


class Schedule:
    def __init__(self):
        self._offset = datetime.timedelta(hours=3)
        self._tz = datetime.timezone(self._offset, name='МСК')
        self._table_name = './static/files/schedule.xlsx'
        self._table = xl.load_workbook(self._table_name)
        self._table_blank = self._table['blank']
        self._animator_to_column = [
            'G', 'I', 'L', 'N', 'Q', 'S', 'V', 'X', 'AA', 'AC', 'AF', 'AH',
            'AK', 'AM', 'AP'
                        'AR'
        ]
        self._animator_salary_to_column = [
            'H', 'J', 'M', 'O', 'R', 'T', 'W', 'Y', 'AB', 'AD', 'AG', 'AI',
            'AL', 'AN', 'AQ', 'AS'
        ]
        self._number_to_char_xl = {
            1: 'H',
            2: 'J',
            3: 'M',
            4: 'O',
            5: 'R',
            6: 'T',
            7: 'W',
            8: 'Y',
            9: 'AB',
            10: 'AD',
            11: 'AG',
            12: 'AI',
            13: 'AL',
            14: 'AN',
            15: 'AQ',
            16: 'AS'
        }
        self._activity_range_xl = range(3, 31)
        self._name_index_xl = 2
        self._starttime_to_row_table = {
            datetime.time(9, 0): '3',
            datetime.time(9, 5): '3',
            datetime.time(9, 10): '3',
            datetime.time(9, 15): '3',
            datetime.time(9, 20): '3',
            datetime.time(9, 25): '3',
            datetime.time(9, 30): '4',
            datetime.time(9, 35): '4',
            datetime.time(9, 40): '4',
            datetime.time(9, 45): '4',
            datetime.time(9, 50): '4',
            datetime.time(9, 55): '4',
            datetime.time(10, 0): '5',
            datetime.time(10, 5): '5',
            datetime.time(10, 10): '5',
            datetime.time(10, 15): '5',
            datetime.time(10, 20): '5',
            datetime.time(10, 25): '5',
            datetime.time(10, 30): '6',
            datetime.time(10, 35): '6',
            datetime.time(10, 40): '6',
            datetime.time(10, 45): '6',
            datetime.time(10, 50): '6',
            datetime.time(10, 55): '6',
            datetime.time(11, 0): '7',
            datetime.time(11, 5): '7',
            datetime.time(11, 10): '7',
            datetime.time(11, 15): '7',
            datetime.time(11, 20): '7',
            datetime.time(11, 25): '7',
            datetime.time(11, 30): '8',
            datetime.time(11, 35): '8',
            datetime.time(11, 40): '8',
            datetime.time(11, 45): '8',
            datetime.time(11, 50): '8',
            datetime.time(11, 55): '8',
            datetime.time(12, 0): '9',
            datetime.time(12, 5): '9',
            datetime.time(12, 10): '9',
            datetime.time(12, 15): '9',
            datetime.time(12, 20): '9',
            datetime.time(12, 25): '9',
            datetime.time(12, 30): '10',
            datetime.time(12, 35): '10',
            datetime.time(12, 40): '10',
            datetime.time(12, 45): '10',
            datetime.time(12, 50): '10',
            datetime.time(12, 55): '10',
            datetime.time(13, 0): '11',
            datetime.time(13, 5): '11',
            datetime.time(13, 10): '11',
            datetime.time(13, 15): '11',
            datetime.time(13, 20): '11',
            datetime.time(13, 25): '11',
            datetime.time(13, 30): '12',
            datetime.time(13, 35): '12',
            datetime.time(13, 40): '12',
            datetime.time(13, 45): '12',
            datetime.time(13, 50): '12',
            datetime.time(13, 55): '12',
            datetime.time(14, 0): '13',
            datetime.time(14, 5): '13',
            datetime.time(14, 10): '13',
            datetime.time(14, 15): '13',
            datetime.time(14, 20): '13',
            datetime.time(14, 25): '13',
            datetime.time(14, 30): '14',
            datetime.time(14, 35): '14',
            datetime.time(14, 40): '14',
            datetime.time(14, 45): '14',
            datetime.time(14, 50): '14',
            datetime.time(14, 55): '14',
            datetime.time(15, 0): '15',
            datetime.time(15, 5): '15',
            datetime.time(15, 10): '15',
            datetime.time(15, 15): '15',
            datetime.time(15, 20): '15',
            datetime.time(15, 25): '15',
            datetime.time(15, 30): '16',
            datetime.time(15, 35): '16',
            datetime.time(15, 40): '16',
            datetime.time(15, 45): '16',
            datetime.time(15, 50): '16',
            datetime.time(15, 55): '16',
            datetime.time(16, 0): '17',
            datetime.time(16, 5): '17',
            datetime.time(16, 10): '17',
            datetime.time(16, 15): '17',
            datetime.time(16, 20): '17',
            datetime.time(16, 25): '17',
            datetime.time(16, 30): '18',
            datetime.time(16, 35): '18',
            datetime.time(16, 40): '18',
            datetime.time(16, 45): '18',
            datetime.time(16, 50): '18',
            datetime.time(16, 55): '18',
            datetime.time(17, 0): '19',
            datetime.time(17, 5): '19',
            datetime.time(17, 10): '19',
            datetime.time(17, 15): '19',
            datetime.time(17, 20): '19',
            datetime.time(17, 25): '19',
            datetime.time(17, 30): '20',
            datetime.time(17, 35): '20',
            datetime.time(17, 40): '20',
            datetime.time(17, 45): '20',
            datetime.time(17, 50): '20',
            datetime.time(17, 55): '20',
            datetime.time(18, 0): '21',
            datetime.time(18, 5): '21',
            datetime.time(18, 10): '21',
            datetime.time(18, 15): '21',
            datetime.time(18, 20): '21',
            datetime.time(18, 25): '21',
            datetime.time(18, 30): '22',
            datetime.time(18, 35): '22',
            datetime.time(18, 40): '22',
            datetime.time(18, 45): '22',
            datetime.time(18, 50): '22',
            datetime.time(18, 55): '22',
            datetime.time(19, 0): '23',
            datetime.time(19, 5): '23',
            datetime.time(19, 10): '23',
            datetime.time(19, 15): '23',
            datetime.time(19, 20): '23',
            datetime.time(19, 25): '23',
            datetime.time(19, 30): '24',
            datetime.time(19, 35): '24',
            datetime.time(19, 40): '24',
            datetime.time(19, 45): '24',
            datetime.time(19, 50): '24',
            datetime.time(19, 55): '24',
            datetime.time(20, 0): '25',
            datetime.time(20, 5): '25',
            datetime.time(20, 10): '25',
            datetime.time(20, 15): '25',
            datetime.time(20, 20): '25',
            datetime.time(20, 25): '25',
            datetime.time(20, 30): '26',
            datetime.time(20, 35): '26',
            datetime.time(20, 40): '26',
            datetime.time(20, 45): '26',
            datetime.time(20, 50): '26',
            datetime.time(20, 55): '26',
            datetime.time(21, 0): '27',
            datetime.time(21, 5): '27',
            datetime.time(21, 10): '27',
            datetime.time(21, 15): '27',
            datetime.time(21, 20): '27',
            datetime.time(21, 25): '27',
            datetime.time(21, 30): '28',
            datetime.time(21, 35): '28',
            datetime.time(21, 40): '28',
            datetime.time(21, 45): '28',
            datetime.time(21, 50): '28',
            datetime.time(21, 55): '28',
            datetime.time(22, 0): '29',
            datetime.time(22, 5): '29',
            datetime.time(22, 10): '29',
            datetime.time(22, 15): '29',
            datetime.time(22, 20): '29',
            datetime.time(22, 25): '29',
            datetime.time(22, 30): '30',
            datetime.time(22, 35): '30',
            datetime.time(22, 40): '30',
            datetime.time(22, 45): '30',
            datetime.time(22, 50): '30',
            datetime.time(22, 55): '30'
        }
        self._full_to_short = {
            'Секрет Выручай-комнаты': 'СВК',
            'Тайна "Летучего Голландца"': 'ТЛГ',
            'Супергерои и перчатка бесконечности': 'СПБ',
            'План побега': 'ПП',
            'Новое дело Шерлока Холмса': 'ШХ',
            'Космический шаттл №5': 'КШ5',
            'Планета Х': 'ПЛХ',
            'Прятки 2.0': '2.0',
            'Прятки 2.0 (аниматор)': '2.0 аним',
            'Прятки 2.0 Kids': '2.0',
            'Прятки 2.0 Kids (аниматор)': '2.0 аним',
            'Прятки 4.0': '4.0',
            'Прятки 4.0 (аниматор)': '4.0 аним',
            'Прятки 4.0 Kids': '4.0',
            'Зеркальные прятки': '4.0',
            'Зеркальные прятки (аниматор)': '4.0 аним',
            'Зеркальные прятки Kids': '4.0',
            'Зеркальные прятки Kids (аниматор)': '4.0 аним',
            'Лазертаг в темноте': 'Лазер',
            'Лазертаг в темноте (аниматор)': 'Лазер аним',
            'Прятки Перфоманс': 'Пр Перф',
            'Прятки Перфоманс (аниматор)': 'Пр Перф аним',
            'Among Us': '2.0 Among Us',
            'Among Us (аниматор)': '2.0 Among аним',
            'Among Us 4.0': '4.0 Among Us',
            'Among Us 4.0 (аниматор)': '4.0 Among Us аним',
            'Игра в кальмара': 'кальмар 2.0',
            'Игра в кальмара (аниматор)': 'кальмар 2.0 аним',
            'Игра в кальмара 4.0': 'кальмар 4.0',
            'Игра в кальмара 4.0 (аниматор)': 'кальмар 4.0 аним',
            'Малый конференц': 'МКз',
            'Большой конференц': 'БКз',
            'Полный конференц': 'ПКз',
            'Космический парк': 'ПР',
            'ВИП-зал': 'ВИП',
            'Арт-зона': 'АРТ',
            'Аниматор в кафе': 'Аним в кафе'
        }
        self._quests = {
            'Секрет Выручай-комнаты', 'Космический шаттл №5',
            'План побега', 'Новое дело Шерлока Холмса',
            'Тайна "Летучего Голландца"', 'Планета Х',
            'Супергерои и перчатка бесконечности'
        }
        self._zones = {
            'Большой конференц', 'Космический парк', 'ВИП-зал', 'Арт-зона',
            'Малый конференц', 'Полный конференц'
        }
        self._activegames = {
            'Лазертаг в темноте', 'Игра в кальмара', 'Прятки 2.0',
            'Among Us', 'Прятки 4.0', 'Игра в кальмара 4.0',
            'Among Us 4.0', 'Прятки Перфоманс', 'Зеркальные прятки Kids',
            'Прятки 2.0 Kids', 'Зеркальные прятки'
        }
        self._games_with_helpers = {
            'Игра в кальмара', 'Among Us', 'Игра в кальмара 4.0',
            'Among Us 4.0', 'Прятки Перфоманс'
        }
        self._anim_games = {
            'Прятки 2.0 (аниматор)', 'Прятки 2.0 Kids (аниматор)',
            'Прятки 4.0 (аниматор)', 'Зеркальные прятки (аниматор)',
            'Зеркальные прятки Kids (аниматор)',
            'Лазертаг в темноте (аниматор)', 'Прятки Перфоманс (аниматор)',
            'Among Us (аниматор)', 'Among Us 4.0 (аниматор)',
            'Игра в кальмара (аниматор)', 'Игра в кальмара 4.0 (аниматор)',
            'Аниматор в кафе'
        }
        self._name_to_color_table = {
            'Секрет Выручай-комнаты': 'E7B8B0',
            'Тайна "Летучего Голландца"': 'FBE5CD',
            'Супергерои и перчатка бесконечности': 'DAEAD0',
            'План побега': 'D9D9D9',
            'Новое дело Шерлока Холмса': 'C9DAF8',
            'Космический шаттл №5': 'D9D1E8',
            'Планета Х': 'DAD2E9',
            'Прятки 2.0': 'F9CA9C',
            'Прятки 2.0 (аниматор)': 'F9CA8C',
            'Прятки 2.0 Kids': 'F9CA9C',
            'Прятки 2.0 Kids (аниматор)': 'F9CA8C',
            'Прятки 4.0': 'B6D7A8',
            'Прятки 4.0 (аниматор)': 'B6D7A9',
            'Прятки 4.0 Kids': 'B6D7A8',
            'Зеркальные прятки': 'B6D7A8',
            'Зеркальные прятки (аниматор)': 'B6D7A9',
            'Зеркальные прятки Kids': 'B6D7A8',
            'Зеркальные прятки Kids (аниматор)': 'B6D7A9',
            'Лазертаг в темноте': 'D6A6BE',
            'Лазертаг в темноте (аниматор)': 'D6A7BE',
            'Прятки Перфоманс': '999999',
            'Прятки Перфоманс (аниматор)': '999998',
            'Among Us': 'F9CA7C',
            'Among Us (аниматор)': 'F9CA6C',
            'Among Us 4.0': 'B6D7A7',
            'Among Us 4.0 (аниматор)': 'B6D7A6',
            'Игра в кальмара': 'EBD446',
            'Игра в кальмара (аниматор)': 'EBD447',
            'Игра в кальмара 4.0': '52EBD7',
            'Игра в кальмара 4.0 (аниматор)': '52EBD8',
            'Малый конференц': 'FFFF00',
            'Большой конференц': '4A86E8',
            'Полный конференц': '69A84F',
            'Космический парк': 'FE0000',
            'ВИП-зал': '9A00FF',
            'Арт-зона': 'FF7FE0',
            'Аниматор в кафе': 'F5B26B'
        }
        self._number_as_hour = {
            1: '9:0-9:30',
            2: '9:30-10:0',
            3: '10:0-10:30',
            4: '10:30-11:0',
            5: '11:0-11:30',
            6: '11:30-12:0',
            7: '12:0-12:30',
            8: '12:30-13:0',
            9: '13:0-13:30',
            10: '13:30-14:0',
            11: '14:0-14:30',
            12: '14:30-15:0',
            13: '15:0-15:30',
            14: '15:30-16:0',
            15: '16:0-16:30',
            16: '16:30-17:0',
            17: '17:0-17:30',
            18: '17:30-18:0',
            19: '18:0-18:30',
            20: '18:30-19:0',
            21: '19:0-19:30',
            22: '19:30-20:0',
            23: '20:0-20:30',
            24: '20:30-21:0',
            25: '21:0-21:30',
            26: '21:30-22:0',
            27: '22:0-22:30',
            28: '22:30-23:0'
        }
        self._activities_number_in_periods = {
            1: 0,
            2: 0,
            3: 0,
            4: 0,
            5: 0,
            6: 0,
            7: 0,
            8: 0,
            9: 0,
            10: 0,
            11: 0,
            12: 0,
            13: 0,
            14: 0,
            15: 0,
            16: 0,
            17: 0,
            18: 0,
            19: 0,
            20: 0,
            21: 0,
            22: 0,
            23: 0,
            24: 0,
            25: 0,
            26: 0,
            27: 0,
            28: 0
        }
        self._grey_color_hex_table = '808080'
        self._animators = None
        self._date = None
        self._active_sheet = None

    def _multisort(self, xs, specs):
        for key, reverse in reversed(specs):
            xs.sort(key=attrgetter(key), reverse=reverse)
        return xs

    def _sort_table_for_weekends(self, lufted_animators, future_activities,
                                 num):
        lufted_start = copy.deepcopy(lufted_animators)
        future_activities_copy = copy.deepcopy(future_activities)
        future_activities_copy = self._multisort(future_activities_copy,
                                                 (('price', True),
                                                  ('start', False)))
        new_animators = lufted_animators
        i = 0
        while i < len(future_activities_copy):
            now_act = len(future_activities_copy)
            new_animators = sorted(
                new_animators,
                key=lambda anim: sum(act.price for act in anim.activities))
            try:
                targetPrice = future_activities_copy[1].price
            except:
                targetPrice = future_activities_copy[0].price
            for animator in new_animators:
                is_ok = False
                currentPrice = 0
                for j in range(len(future_activities_copy)):
                    if j >= len(future_activities_copy):
                        break
                    cnt_of_unavailable_activities = 0
                    for activity in animator.activities:
                        if isinstance(activity, LuftTime):
                            activity_start = activity.start
                            activity_finish = activity.finish
                        else:
                            activity_start = activity.start.time()
                            activity_finish = activity.finish.time()
                        future_activity_start = future_activities_copy[
                            j].start.time()
                        future_activity_finish = future_activities_copy[
                            j].finish.time()
                        if (activity_start <= future_activity_start <
                            activity_finish) or (
                                activity_start < future_activity_finish <
                                activity_finish) or (
                                future_activity_start <= activity_start
                                < future_activity_finish):
                            cnt_of_unavailable_activities += 1
                    if cnt_of_unavailable_activities == 0:
                        animator.activities.append(future_activities_copy[j])
                        currentPrice += future_activities_copy[j].price
                        future_activities_copy.remove(
                            future_activities_copy[j])
                        if currentPrice >= targetPrice:
                            is_ok = True
                            break
                    else:
                        continue
                if is_ok:
                    break
            if len(future_activities_copy) == now_act:
                num += 1
                lufted_start.append(Animator(name=f'{num}'))
                return self._sort_table_for_weekends(lufted_start,
                                                     future_activities, num)
        for animator in new_animators:
            animator.total_salary_reload()
        return new_animators

    def _sort_table(self, lufted_animators, future_activities, num):
        lufted_start = copy.deepcopy(lufted_animators)
        future_activities_copy = copy.deepcopy(future_activities)
        future_activities_copy = self._multisort(future_activities_copy,
                                                 (('start', False),
                                                  ('price', True)))
        new_animators = lufted_animators
        i = 0
        try:
            targetPrice = future_activities_copy[1].price
        except:
            targetPrice = future_activities_copy[0].price
        while i < len(future_activities_copy):
            now_act = len(future_activities_copy)
            new_animators = sorted(
                new_animators,
                key=lambda anim: sum(act.price for act in anim.activities))
            for animator in new_animators:
                is_ok = False
                currentPrice = 0
                for j in range(len(future_activities_copy)):
                    if j >= len(future_activities_copy):
                        break
                    cnt_of_unavailable_activities = 0
                    for activity in animator.activities:
                        if isinstance(activity, LuftTime):
                            activity_start = activity.start
                            activity_finish = activity.finish
                        else:
                            activity_start = activity.start.time()
                            activity_finish = activity.finish.time()
                        future_activity_start = future_activities_copy[
                            j].start.time()
                        future_activity_finish = future_activities_copy[
                            j].finish.time()
                        if (activity_start <= future_activity_start <
                            activity_finish) or (
                                activity_start < future_activity_finish <
                                activity_finish) or (
                                future_activity_start <= activity_start
                                < future_activity_finish):
                            cnt_of_unavailable_activities += 1
                    if cnt_of_unavailable_activities == 0:
                        animator.activities.append(future_activities_copy[j])
                        currentPrice += future_activities_copy[j].price
                        future_activities_copy.remove(
                            future_activities_copy[j])
                        if currentPrice >= targetPrice - 500:
                            is_ok = True
                            targetPrice = currentPrice
                            break
                    else:
                        continue
                if is_ok:
                    break
            if len(future_activities_copy) == now_act:
                num += 1
                lufted_start.append(Animator(name=f'{num}'))
                return self._sort_table(lufted_start, future_activities, num)
        for animator in new_animators:
            animator.total_salary_reload()
        return new_animators

    def _get_salary(self, current_activities):
        return sum(current_activities[i].price for i in range(len(current_activities)))

    def _activities_number_in_periods_reload(self):
        self._activities_number_in_periods = {
            1: 0,
            2: 0,
            3: 0,
            4: 0,
            5: 0,
            6: 0,
            7: 0,
            8: 0,
            9: 0,
            10: 0,
            11: 0,
            12: 0,
            13: 0,
            14: 0,
            15: 0,
            16: 0,
            17: 0,
            18: 0,
            19: 0,
            20: 0,
            21: 0,
            22: 0,
            23: 0,
            24: 0,
            25: 0,
            26: 0,
            27: 0,
            28: 0
        }

    def _get_table(self, lufted_animators, future_activities):
        animators = lufted_animators
        for i in range(len(future_activities)):
            cnt_unavailable_animators = 0
            for j in range(len(animators)):
                for k in range(len(animators[j].activities)):
                    if isinstance(animators[j].activities[k], LuftTime):
                        activity_start = animators[j].activities[k].start
                        activity_finish = animators[j].activities[k].finish
                    else:
                        activity_start = animators[j].activities[k].start.time(
                        )
                        activity_finish = animators[j].activities[
                            k].finish.time()
                    if isinstance(future_activities[i], LuftTime):
                        future_activity_start = future_activities[i].start
                        future_activity_finish = future_activities[i].finish
                    else:
                        future_activity_start = future_activities[
                            i].start.time()
                        future_activity_finish = future_activities[
                            i].finish.time()
                    if (activity_start <= future_activity_start < activity_finish) \
                            or (activity_start < future_activity_finish <= activity_finish) \
                            or (future_activity_start <= activity_start < future_activity_finish):
                        cnt_unavailable_animators = cnt_unavailable_animators + 1
                        break
                else:
                    animators[j].activities.append(future_activities[i])
                    break
            if cnt_unavailable_animators == len(animators):
                animators.append(Animator())
                animators[-1].activities.append(future_activities[i])
        return animators

    def _try_weekend_sort(self, current_animators, current_activities):
        animators_cnt = len(current_animators)
        activities_copy = copy.deepcopy(current_activities)
        anims1 = self._sort_table(copy.deepcopy(current_animators),
                                  copy.deepcopy(current_activities),
                                  len(current_animators))
        anims2 = self._sort_table_for_weekends(
            copy.deepcopy(current_animators),
            copy.deepcopy(current_activities), len(current_animators))
        minn1 = min(len(anims1), len(anims2))
        if animators_cnt < minn1:
            for act in activities_copy:
                start = datetime.time(act.start.hour, act.start.minute)
                finish = datetime.time(act.finish.hour, act.finish.minute)
                for i in range(len(self._number_as_hour)):
                    per_start = datetime.time(
                        int(self._number_as_hour[i + 1].split('-')[0].split(
                            ':')[0]),
                        int(self._number_as_hour[i + 1].split('-')[0].split(
                            ':')[1]))
                    per_finish = datetime.time(
                        int(self._number_as_hour[i + 1].split('-')[1].split(
                            ':')[0]),
                        int(self._number_as_hour[i + 1].split('-')[1].split(
                            ':')[1]))
                    if (start < per_start <
                        finish) or (start < per_finish < finish) or (
                            per_start < finish <= per_finish) or (
                            per_start < start < per_finish):
                        if isinstance(act, Helper):
                            self._activities_number_in_periods[i + 1] += 0.51
                        else:
                            self._activities_number_in_periods[i + 1] += 1
            end = False
            mmax = -1
            while True:
                max_hours = max(self._activities_number_in_periods.values())
                if max_hours == mmax:
                    break
                mmax = max_hours
                max_hours_inds = []
                for i in range(len(self._activities_number_in_periods)):
                    if self._activities_number_in_periods[i + 1] == max_hours:
                        max_hours_inds.append(i + 1)
                for i in max_hours_inds:
                    if isinstance(self._activities_number_in_periods[i],
                                  float):
                        period_start = datetime.time(
                            int(self._number_as_hour[i].split('-')[0].split(
                                ':')[0]),
                            int(self._number_as_hour[i].split('-')[0].split(
                                ':')[1]))
                        period_finish = datetime.time(
                            int(self._number_as_hour[i].split('-')[1].split(
                                ':')[0]),
                            int(self._number_as_hour[i].split('-')[1].split(
                                ':')[1]))
                        for act in activities_copy:
                            startt = datetime.time(act.start.hour,
                                                   act.start.minute)
                            finishh = datetime.time(act.finish.hour,
                                                    act.finish.minute)
                            if isinstance(
                                    act, Helper
                            ) and act.name != 'Прятки Перфоманс' and (
                                    (startt < period_start < finishh) or
                                    (startt < period_finish < finishh) or
                                    (period_start < finishh <= period_finish) or
                                    (period_start < startt < period_finish)):
                                activities_copy.remove(act)
                                self._activities_number_in_periods[i] -= 0.5
                                break
                    else:
                        end = True
                        break
                if end:
                    break
            anims3 = self._sort_table(copy.deepcopy(current_animators),
                                      activities_copy, len(current_animators))
            anims4 = self._sort_table_for_weekends(
                copy.deepcopy(current_animators), activities_copy,
                len(current_animators))
            minn2 = min(len(anims3), len(anims4))
            if minn2 < minn1:
                anims1 = anims3
                anims2 = anims4
        self._activities_number_in_periods_reload()
        if not anims1:
            return []
        return anims2

    def _get_activities(self, active_games, corps):
        future_activities = []
        for g in active_games:
            if g.get('name') in self._activegames:
                try:
                    cnt_children = g.get('number of children')
                except:
                    cnt_children = 1
                have_helper = False
                if cnt_children // 15 > 0:
                    have_helper = True
                future_activities.append(
                    Game(
                        g.get('name'),
                        datetime.datetime.strptime(
                            f'{g.get("date").hour}:{g.get("date").minute}',
                            '%H:%M'),
                        datetime.datetime.strptime(
                            f'{g.get("date").hour + 1}:{g.get("date").minute}',
                            '%H:%M')))
                if g.get('animator'):
                    future_activities.append(
                        GameAnim(
                            f'{g.get("name")} (аниматор)',
                            datetime.datetime.strptime(
                                f'{g.get("date").hour}:{g.get("date").minute}',
                                '%H:%M'),
                            datetime.datetime.strptime(
                                f'{g.get("date").hour + 1}:{g.get("date").minute}',
                                '%H:%M')))
                    have_helper = True
                if g.get('name'
                         ) in self._games_with_helpers and not have_helper:
                    if g.get('name') == 'Among Us' or g.get(
                            'name') == 'Among Us 4.0':
                        future_activities.append(
                            Helper(
                                g.get('name'),
                                datetime.datetime.strptime(
                                    f'{g.get("date").hour}:{g.get("date").minute}',
                                    '%H:%M'),
                                datetime.datetime.strptime(
                                    f'{(g.get("date") + datetime.timedelta(minutes=30)).hour}:'
                                    f'{(g.get("date") + datetime.timedelta(minutes=30)).minute}',
                                    '%H:%M')))
                    else:
                        future_activities.append(
                            Helper(
                                g.get('name'),
                                datetime.datetime.strptime(
                                    f'{(g.get("date") + datetime.timedelta(minutes=30)).hour}:'
                                    f'{(g.get("date") + datetime.timedelta(minutes=30)).minute}',
                                    '%H:%M'),
                                datetime.datetime.strptime(
                                    f'{g.get("date").hour + 1}:{g.get("date").minute}',
                                    '%H:%M')))
            elif g.get('name') in self._quests and g.get('animator'):
                future_activities.append(
                    Quest(
                        g.get('name'),
                        datetime.datetime.strptime(
                            f'{g.get("date").hour}:{g.get("date").minute}',
                            '%H:%M'),
                        datetime.datetime.strptime(
                            f'{g.get("date").hour + 1}:{g.get("date").minute}',
                            '%H:%M')))
        for c in corps:
            try:
                cnt_children = c.get('number of children')
            except:
                cnt_children = 1
            if cnt_children is None:
                cnt_children = 1
            mk_is_added = False
            for z in c.get('zones'):
                start_time = datetime.datetime.strptime(
                    z.get('time').split(' - ')[0], '%H:%M')
                finish_time = datetime.datetime.strptime(
                    z.get('time').split(' - ')[1], '%H:%M')
                if (finish_time - start_time).seconds < 1800:
                    continue
                time_helper = [(start_time, finish_time)]
                add_services = 0
                games = 0
                mk = 0
                add_services_starts = set()
                for ad_s in c.get('addservices'):
                    if (ad_s.get('time') != '00:00'
                        and ad_s.get('name') != 'Воздушные шары' and
                        (start_time <= datetime.datetime.strptime(
                            ad_s.get('time_start'), '%H:%M') < finish_time)
                        and ad_s.get('time') not in add_services_starts
                    ) or ad_s.get('name') == 'Мастер-класс':
                        start = datetime.datetime(1990, 1, 1, 0, 0)
                        finish = datetime.datetime(1990, 1, 1, 0, 0)
                        if ad_s.get('name') == 'Мастер-класс' and ad_s.get(
                                'time_start') != '00:00':
                            start = datetime.datetime.strptime(
                                ad_s.get('time').split(' - ')[0], '%H:%M')
                            finish = datetime.datetime.strptime(
                                ad_s.get('time').split(' - ')[1], '%H:%M')
                            if start_time <= datetime.datetime.strptime(
                                    ad_s.get('time_start'),
                                    '%H:%M') < finish_time:
                                mk += (finish - start).seconds / 3600
                                add_services_starts.add(ad_s.get('time'))
                        elif ad_s.get('name') == 'Мастер-класс' and ad_s.get(
                                'time_start') == '00:00' and not mk_is_added:
                            mk += 1
                            mk_is_added = True
                            add_services_starts.add(ad_s.get('time'))
                        elif ad_s.get('name') != 'Мастер-класс' \
                                and ad_s.get('name') != 'Фотограф' \
                                and ad_s.get('name') != 'Видеограф' \
                                and ad_s.get('name') != 'Дополнительный игрок' \
                                and ad_s.get('name') != 'Кенди-бар' \
                                and ad_s.get('name') != 'Костюмированный аниматор' \
                                and ad_s.get('name') != 'Аниматор в квест' \
                                and ad_s.get('name') != 'Воздушные шары' \
                                and ad_s.get('name') != 'Фотозона' \
                                and ad_s.get('name') != 'ИЦ закрыт':
                            try:
                                start = datetime.datetime.strptime(
                                    ad_s.get('time').split(' - ')[0], '%H:%M')
                                finish = datetime.datetime.strptime(
                                    ad_s.get('time').split(' - ')[1], '%H:%M')
                                add_services += (finish - start).seconds / 3600
                                add_services_starts.add(ad_s.get('time'))
                            except:
                                pass
                        for i in range(len(time_helper)):
                            th_start = time_helper[i][0]
                            th_finish = time_helper[i][1]
                            if th_start < start and th_finish > finish:
                                new_tuple_1 = (th_start, start)
                                new_tuple_2 = (finish, th_finish)
                                time_helper.insert(i, new_tuple_1)
                                time_helper.insert(i + 1, new_tuple_2)
                                time_helper.pop(i + 2)
                                break
                            elif th_start <= start < th_finish and th_finish > finish:
                                new_tuple = (finish, th_finish)
                                time_helper.insert(i, new_tuple)
                                time_helper.pop(i + 1)
                                break
                            elif th_start < start and th_finish == finish:
                                new_tuple = (th_start, start)
                                time_helper.insert(i, new_tuple)
                                time_helper.pop(i + 1)
                                break
                            elif th_start == start and th_finish == finish:
                                time_helper.pop(i)
                                break
                games_starts = set()
                for g in c.get('games'):
                    start = datetime.datetime.strptime(g.get('time'), '%H:%M')
                    finish = start + datetime.timedelta(hours=1)
                    if (start_time <= datetime.datetime.strptime(
                            g.get('time'), '%H:%M') <
                        finish_time) and g.get('time') not in games_starts:
                        games += 1
                        games_starts.add(g.get('time'))
                        for i in range(len(time_helper)):
                            th_start = time_helper[i][0]
                            th_finish = time_helper[i][1]
                            if th_start < start and th_finish > finish:
                                new_tuple_1 = (th_start, start)
                                new_tuple_2 = (finish, th_finish)
                                time_helper.insert(i, new_tuple_1)
                                time_helper.insert(i + 1, new_tuple_2)
                                time_helper.pop(i + 2)
                                break
                            elif th_start <= start < th_finish and th_finish > finish:
                                new_tuple = (finish, th_finish)
                                time_helper.insert(i, new_tuple)
                                time_helper.pop(i + 1)
                                break
                            elif th_start < start and th_finish == finish:
                                new_tuple = (th_start, start)
                                time_helper.insert(i, new_tuple)
                                time_helper.pop(i + 1)
                                break
                            elif th_start == start and th_finish == finish:
                                time_helper.pop(i)
                                break
                future_activities.append(
                    Zone(z.get('name'), start_time, finish_time, add_services,
                         games, mk, z.get('table')))
                if cnt_children >= 15:
                    for th in time_helper:
                        if (th[1] - th[0]).seconds >= 1800:
                            future_activities.append(
                                Zone(z.get('name'), th[0], th[1], 0, 0, 0,
                                     z.get('table')))
            if cnt_children >= 15:
                for g in c.get('games'):
                    if g.get('name') in self._activegames:
                        for f_a in future_activities:
                            start = datetime.datetime.strptime(
                                g.get('time'), '%H:%M')
                            finish = datetime.datetime.strptime(
                                f'{datetime.datetime.strptime(g.get("time"), "%H:%M").hour + 1}:{datetime.datetime.strptime(g.get("time"), "%H:%M").minute}',
                                '%H:%M')
                            if isinstance(
                                    f_a,
                                    Helper) and f_a.name == g.get('name') and (
                                    start <= f_a.start <= finish
                            ) and (start_time <= start <= finish_time):
                                future_activities.remove(f_a)
                                future_activities.append(
                                    Game(
                                        g.get('name'), start,
                                        datetime.datetime.strptime(
                                            f'{start.hour + 1}:{start.minute}',
                                            '%H:%M')))
                                break
                            elif isinstance(
                                    f_a, Game
                            ) and f_a.name not in self._games_with_helpers and f_a.name == g.get(
                                'name') and start == f_a.start:
                                future_activities.append(
                                    Game(
                                        g.get('name'), start,
                                        datetime.datetime.strptime(
                                            f'{start.hour + 1}:{start.minute}',
                                            '%H:%M')))
                                break
            for ad_s in c.get('addservices'):
                if ad_s.get('name') == 'Аниматор в кафе' or ad_s.get(
                        'name') == 'Аниматор няня':
                    start = datetime.datetime.strptime(
                        ad_s.get('time').split(' - ')[0], '%H:%M')
                    finish = datetime.datetime.strptime(
                        ad_s.get('time').split(' - ')[1], '%H:%M')
                    future_activities.append(
                        Quest(ad_s.get('name'), start, finish))
        self._zones_round_up(future_activities)
        self._add_colour_info(future_activities)
        self._add_cell_text_for_activities(future_activities)
        return future_activities

    def _add_cell_text_for_activities(self, activities):
        for act in activities:
            if act.cell_text == '':
                act.cell_text = self._full_to_short.get(act.name)

    def _add_colour_info(self, activities):
        for act in activities:
            act.colour_info = self._name_to_color_table.get(act.name)

    def _zones_round_up(self, activities):
        for act in activities:
            if isinstance(act, Zone):
                if act.hours_200 < 0:
                    act.hours_200 = 0
                elif 0 < act.hours_200 < 0.75:
                    act.hours_200 = 0.5
                elif 0.75 <= act.hours_200 < 1.25:
                    act.hours_200 = 1
                elif 1.25 <= act.hours_200 < 1.75:
                    act.hours_200 = 1.5
                elif 1.75 <= act.hours_200 < 2.25:
                    act.hours_200 = 2
                elif 2.25 <= act.hours_200 < 2.75:
                    act.hours_200 = 2.5
                elif 2.75 <= act.hours_200 < 3.25:
                    act.hours_200 = 3
                elif 3.25 <= act.hours_200 < 3.75:
                    act.hours_200 = 3.5
                elif 3.75 <= act.hours_200 < 4.25:
                    act.hours_200 = 4
                elif 4.25 <= act.hours_200 < 4.75:
                    act.hours_200 = 4.5
                elif 4.75 <= act.hours_200 < 5.25:
                    act.hours_200 = 5
                if act.hours_350 < 0:
                    act.hours_350 = 0
                elif 0 < act.hours_350 < 0.75:
                    act.hours_350 = 0.5
                elif 0.75 <= act.hours_350 < 1.25:
                    act.hours_350 = 1
                elif 1.25 <= act.hours_350 < 1.75:
                    act.hours_350 = 1.5
                elif 1.75 <= act.hours_350 < 2.25:
                    act.hours_350 = 2
                elif 2.25 <= act.hours_350 < 2.75:
                    act.hours_350 = 2.5
                elif 2.75 <= act.hours_350 < 3.25:
                    act.hours_350 = 3
                elif 3.25 <= act.hours_350 < 3.75:
                    act.hours_350 = 3.5
                elif 3.75 <= act.hours_350 < 4.25:
                    act.hours_350 = 4
                elif 4.25 <= act.hours_350 < 4.75:
                    act.hours_350 = 4.5
                elif 4.75 <= act.hours_350 < 5.25:
                    act.hours_350 = 5
                if act.hours_500 < 0:
                    act.hours_500 = 0
                elif 0 < act.hours_500 < 0.75:
                    act.hours_500 = 0.5
                elif 0.75 <= act.hours_500 < 1.25:
                    act.hours_500 = 1
                elif 1.25 <= act.hours_500 < 1.75:
                    act.hours_500 = 1.5
                elif 1.75 <= act.hours_500 < 2.25:
                    act.hours_500 = 2
                elif 2.25 <= act.hours_500 < 2.75:
                    act.hours_500 = 2.5
                elif 2.75 <= act.hours_500 < 3.25:
                    act.hours_500 = 3
                elif 3.25 <= act.hours_500 < 3.75:
                    act.hours_500 = 3.5
                elif 3.75 <= act.hours_500 < 4.25:
                    act.hours_500 = 4
                elif 4.25 <= act.hours_500 < 4.75:
                    act.hours_500 = 4.5
                elif 4.75 <= act.hours_500 < 5.25:
                    act.hours_500 = 5
        return activities

    def _get_animators_weekdays(self, date):
        crm_games, crm_corps = CRM().get_activities(date)
        activities = self._get_activities(crm_games, crm_corps)
        if len(activities) == 0:
            return []
        activities_copy = copy.deepcopy(activities)
        anims1 = self._sort_table([], activities, 0)
        anims2 = self._sort_table_for_weekends([], activities, 0)
        minn1 = min(len(anims1), len(anims2))
        for act in activities_copy:
            start = datetime.time(act.start.hour, act.start.minute)
            finish = datetime.time(act.finish.hour, act.finish.minute)
            for i in range(len(self._number_as_hour)):
                per_start = datetime.time(
                    int(self._number_as_hour[i +
                                             1].split('-')[0].split(':')[0]),
                    int(self._number_as_hour[i +
                                             1].split('-')[0].split(':')[1]))
                per_finish = datetime.time(
                    int(self._number_as_hour[i +
                                             1].split('-')[1].split(':')[0]),
                    int(self._number_as_hour[i +
                                             1].split('-')[1].split(':')[1]))
                if (start < per_start <
                    finish) or (start < per_finish < finish) or (
                        per_start < finish <=
                        per_finish) or (per_start < start < per_finish):
                    if isinstance(act, Helper):
                        self._activities_number_in_periods[i + 1] += 0.51
                    else:
                        self._activities_number_in_periods[i + 1] += 1
        end = False
        mmax = -1
        while True:
            max_hours = max(self._activities_number_in_periods.values())
            if max_hours == mmax:
                break
            mmax = max_hours
            max_hours_inds = []
            for i in range(len(self._activities_number_in_periods)):
                if self._activities_number_in_periods[i + 1] == max_hours:
                    max_hours_inds.append(i + 1)
            for i in max_hours_inds:
                if isinstance(self._activities_number_in_periods[i], float):
                    period_start = datetime.time(
                        int(self._number_as_hour[i].split('-')[0].split(':')
                            [0]),
                        int(self._number_as_hour[i].split('-')[0].split(':')
                            [1]))
                    period_finish = datetime.time(
                        int(self._number_as_hour[i].split('-')[1].split(':')
                            [0]),
                        int(self._number_as_hour[i].split('-')[1].split(':')
                            [1]))
                    for act in activities_copy:
                        startt = datetime.time(act.start.hour,
                                               act.start.minute)
                        finishh = datetime.time(act.finish.hour,
                                                act.finish.minute)
                        if isinstance(
                                act, Helper
                        ) and act.name != 'Прятки Перфоманс' and (
                                (startt < period_start < finishh) or
                                (startt < period_finish < finishh) or
                                (period_start < finishh <= period_finish) or
                                (period_start < startt < period_finish)):
                            activities_copy.remove(act)
                            self._activities_number_in_periods[i] -= 0.5
                            break
                else:
                    end = True
                    break
            if end:
                break
        anims3 = self._sort_table([], activities_copy, 0)
        anims4 = self._sort_table_for_weekends([], activities_copy, 0)
        minn2 = min(len(anims3), len(anims4))
        if minn2 < minn1:
            anims1 = anims3
            anims2 = anims4
        self._activities_number_in_periods_reload()
        if not anims1:
            return []
        if len(anims1) > len(anims2):
            return anims2
        elif len(anims2) > len(anims1):
            return anims1
        else:
            return anims2

    def _get_animators_weekends(self, timedelta_info, date):
        crm_games, crm_corps = CRM().get_activities(date)
        animators = list()
        if timedelta_info is not None:
            for i in range(len(timedelta_info)):
                if timedelta_info[i].get('timedelta') is not None:
                    animators.append(
                        Animator(timedelta_info[i].get('name')))
                    start_luft = datetime.datetime.strptime(timedelta_info[i].get('timedelta')[0], '%H:%M').time()
                    finish_luft = datetime.datetime.strptime(timedelta_info[i].get('timedelta')[1], '%H:%M').time()
                    animators[-1].add_luft(start_luft, finish_luft)
        activities = self._get_activities(crm_games, crm_corps)
        return self._try_weekend_sort(animators, activities)

    def _make_new_sheet(self, date):
        self._table = xl.load_workbook(self._table_name)
        if date.strftime('%d.%m.%Y') not in self._table.sheetnames:
            self._table_blank = self._table['blank']
            self._table.create_sheet(date.strftime('%d.%m.%Y'))
            self._table.save(self._table_name)
            self._active_sheet = self._table[date.strftime('%d.%m.%Y')]
            WorksheetCopy(self._table_blank, self._active_sheet).copy_worksheet()
            self._table.save(self._table_name)
            new_sheet = self._table[date.strftime('%d.%m.%Y')]
            return new_sheet
        else:
            raise Exception('На эту дату уже создан график!')

    def _endtime_to_row_table(self, start, finish, start_row):
        if isinstance(start, datetime.datetime) and isinstance(
                finish, datetime.datetime):
            delta_datetime = (finish - start).seconds / 1800
            if delta_datetime < 1:
                delta_datetime = 1
            else:
                delta_datetime = int(delta_datetime)
            return str(int(start_row) + delta_datetime - 1)
        else:
            delta_time = ((finish.hour * 60 + finish.minute) -
                          (start.hour * 60 + start.minute)) / 30
            if delta_time < 1:
                delta_time = 1
            else:
                delta_time = int(delta_time)
            return str(int(start_row) + delta_time - 1)

    def _paste_activities_salary_total_salary_table(self, animators):
        self._table = xl.load_workbook(self._table_name)
        min_hour = 24
        max_hour = 0
        for anim in animators:
            for act in anim.activities:
                if act.start.hour < min_hour:
                    min_hour = act.start.hour
        for anim in animators:
            for act in anim.activities:
                if act.finish.hour > max_hour:
                    max_hour = act.finish.hour
        self._table_blank = self._table['blank']
        self._active_sheet = self._table[self._date.strftime('%d.%m.%Y')]
        for i in range(0, len(animators)):
            for j in range(0, len(animators[i].activities)):
                if not isinstance(animators[i].activities[j], LuftTime):
                    if isinstance(animators[i].activities[j], Zone):
                        self._active_sheet[self._animator_to_column[i] +
                                           self._starttime_to_row_table[animators[i].activities[j].start.time()]
                                           ] = animators[i].activities[j].start.strftime("%H:%M") + '-' + \
                                               animators[i].activities[j].finish.strftime("%H:%M") + \
                                               ' ' + self._full_to_short[animators[i].activities[j].name]
                        if animators[i].activities[j].mk > 0:
                            self._active_sheet[
                                self._animator_to_column[i] +
                                self._starttime_to_row_table[
                                    animators[i].activities[j].start.time(
                                    )]].value += ' мк'
                            animators[i].activities[j].cell_text += ' мк'
                        if (animators[i].activities[j].finish.hour +
                            animators[i].activities[j].finish.minute / 60
                        ) - (animators[i].activities[j].start.hour +
                             animators[i].activities[j].start.minute /
                             60) >= 1:
                            self._active_sheet[
                                self._animator_to_column[i] +
                                self._starttime_to_row_table[
                                    animators[i].activities[j].start.time(
                                    )]].alignment = Alignment(
                                wrap_text=True,
                                horizontal='center',
                                vertical='center')
                        else:
                            self._active_sheet[
                                self._animator_to_column[i] +
                                self._starttime_to_row_table[
                                    animators[i].activities[j].start.time(
                                    )]].alignment = Alignment(
                                wrap_text=True,
                                horizontal='center',
                                vertical='center')
                            self._active_sheet[self._animator_to_column[i] +
                                               self._starttime_to_row_table[
                                                   animators[i].activities[j].
                                                   start.time()]].font = Font(
                                size='11',
                                name='Nunito',
                                bold=True)
                        self._active_sheet[
                            self._animator_to_column[i] +
                            self._starttime_to_row_table[
                                animators[i].activities[j].start.time(
                                )]].fill = PatternFill(
                            'solid',
                            fgColor=Color(rgb=self._name_to_color_table[
                                animators[i].activities[j].name]))
                    else:
                        self._active_sheet[self._animator_to_column[i] +
                                           self._starttime_to_row_table[animators[i].activities[j].start.time()]
                                           ] = animators[i].activities[j].start.strftime("%H:%M") + \
                                               ' ' + self._full_to_short[animators[i].activities[j].name]
                        if (animators[i].activities[j].finish.hour + animators[i].activities[j].finish.minute / 60) - (
                                animators[i].activities[j].start.hour + animators[i].activities[
                            j].start.minute / 60) >= 1:
                            self._active_sheet[self._animator_to_column[i] +
                                               self._starttime_to_row_table[animators[i].activities[j].start.time()]
                                               ].alignment = Alignment(wrap_text=True, horizontal='center',
                                                                       vertical='center')
                        else:
                            self._active_sheet[self._animator_to_column[i] +
                                               self._starttime_to_row_table[animators[i].activities[j].start.time()]
                                               ].alignment = Alignment(wrap_text=True, horizontal='center',
                                                                       vertical='center')
                            self._active_sheet[self._animator_to_column[i] +
                                               self._starttime_to_row_table[animators[i].activities[j].start.time()]
                                               ].font = Font(size='11', name='Nunito', bold=True)
                        self._active_sheet[self._animator_to_column[i] +
                                           self._starttime_to_row_table[animators[i].activities[j].start.time()]
                                           ].fill = PatternFill('solid',
                                                                fgColor=self._name_to_color_table[
                                                                    animators[i].activities[j].name])
                    if isinstance(animators[i].activities[j], Zone):
                        self._active_sheet[
                            self._animator_salary_to_column[i] +
                            self._starttime_to_row_table[
                                animators[i].activities[j].start.time(
                                )]] = f'={animators[i].activities[j].hours_200}*200+{animators[i].activities[j].hours_350}*350+{animators[i].activities[j].hours_500}*500+{animators[i].activities[j].mk}*500'
                    else:
                        self._active_sheet[
                            self._animator_salary_to_column[i] +
                            self._starttime_to_row_table[
                                animators[i].activities[j].start.time(
                                )]] = int(animators[i].activities[j].price)
                    self._table.save(self._table_name)
                    self._active_sheet.merge_cells(
                        self._animator_to_column[i] +
                        self._starttime_to_row_table[
                            animators[i].activities[j].start.time()] + ':' +
                        self._animator_to_column[i] +
                        self._endtime_to_row_table(
                            animators[i].activities[j].start.time(),
                            animators[i].activities[j].finish.time(),
                            self._starttime_to_row_table[
                                animators[i].activities[j].start.time()]))
                    self._table.save(self._table_name)
                    self._active_sheet.merge_cells(
                        self._animator_salary_to_column[i] +
                        self._starttime_to_row_table[
                            animators[i].activities[j].start.time()] + ':' +
                        self._animator_salary_to_column[i] +
                        self._endtime_to_row_table(
                            animators[i].activities[j].start.time(),
                            animators[i].activities[j].finish.time(),
                            self._starttime_to_row_table[
                                animators[i].activities[j].start.time()]))
                    self._table.save(self._table_name)
                else:
                    self._active_sheet[
                        self._animator_to_column[i] +
                        self._starttime_to_row_table[
                            animators[i].activities[j].
                            start]].fill = PatternFill(
                        'solid', fgColor=self._grey_color_hex_table)
                    self._table.save(self._table_name)
                    self._active_sheet.merge_cells(
                        self._animator_to_column[i] +
                        self._starttime_to_row_table[
                            animators[i].activities[j].start] + ':' +
                        self._animator_to_column[i] +
                        self._endtime_to_row_table(
                            animators[i].activities[j].start, animators[i].
                            activities[j].finish, self._starttime_to_row_table[
                                animators[i].activities[j].start]))
                    self._table.save(self._table_name)
                    self._active_sheet.merge_cells(
                        self._animator_salary_to_column[i] +
                        self._starttime_to_row_table[
                            animators[i].activities[j].start] + ':' +
                        self._animator_salary_to_column[i] +
                        self._endtime_to_row_table(
                            animators[i].activities[j].start, animators[i].
                            activities[j].finish, self._starttime_to_row_table[
                                animators[i].activities[j].start]))
        if min_hour >= 11:
            for i in range(3, (min_hour - 11) * 2 + 5):
                self._active_sheet.row_dimensions[i].hidden = True
        if max_hour <= 21:
            for i in range(29 - (21 - max_hour) * 2, 31):
                self._active_sheet.row_dimensions[i].hidden = True
        self._table.save(self._table_name)

    def _paste_name_table(self, name, ind):
        self._table = xl.load_workbook(self._table_name)
        self._table_blank = self._table['blank']
        self._active_sheet = self._table[self._date.strftime('%d.%m.%Y')]
        self._active_sheet[self._animator_to_column[ind] + '2'] = name
        self._table.save(self._table_name)

    def make_schedule_weekdays(self, date):
        self._animators = self._get_animators_weekdays(date)
        self._date = date
        self._active_sheet = self._make_new_sheet(date)
        self._paste_activities_salary_total_salary_table(self._animators)
        return self._table_name

    def make_schedule_weekends(self, date, timedelta_info):
        self._animators = self._get_animators_weekends(timedelta_info, date)
        self._date = date
        self._active_sheet = self._make_new_sheet(date)
        self._paste_activities_salary_total_salary_table(self._animators)
        for i in range(0, len(self._animators)):
            self._paste_name_table(self._animators[i].name, i)
        return self._table_name
